"""
Payment Summary Screen - View Payment Details
"""

import customtkinter as ctk
from tkinter import messagebox
from datetime import date, datetime
import sys
import os
from pathlib import Path
import openpyxl
from openpyxl import Workbook

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))

from app.database.models import SessionLocal, FeeChallan, Student, Guardian, FeeStatus
from app.utils.center_window import center_and_maximize  # UI FIX


class PaymentSummaryScreen(ctk.CTkToplevel):
    """Payment Summary Screen - View Payment Details"""
    
    def __init__(self, parent, db=None):
        super().__init__(parent)
        
        self.title("Payment Summary")
        self.geometry("1600x850")  # Increased to fit more columns
        self.db = db if db else SessionLocal()
        
        # Create reports directory
        self.reports_dir = os.path.join(
            os.path.dirname(os.path.dirname(os.path.dirname(__file__))),
            "reports"
        )
        Path(self.reports_dir).mkdir(exist_ok=True)
        
        # UI FIX: Center and maximize window after creation
        self.after(100, lambda: center_and_maximize(self))
        
        # Create UI
        self.create_widgets()
        self.load_data()
        
        # Make modal
        self.transient(parent)
    
    def refresh_summary(self):
        """Refresh the summary data after a payment is made"""
        self.load_data()
    
    def create_widgets(self):
        """Create main UI widgets"""
        
        # Main container
        main_frame = ctk.CTkFrame(self, fg_color="#f0f2f5", corner_radius=10)
        main_frame.pack(fill="both", expand=True, padx=20, pady=20)
        
        # Title
        title = ctk.CTkLabel(
            main_frame,
            text="Payment Summary",
            font=("Arial", 24, "bold"),
            text_color="#1e3a5f"
        )
        title.pack(pady=20)
        
        # Month and Year Selector
        selector_frame = ctk.CTkFrame(main_frame, fg_color="white", corner_radius=10)
        selector_frame.pack(fill="x", padx=20, pady=10)
        
        selector_label = ctk.CTkLabel(
            selector_frame,
            text="Select Month and Year:",
            font=("Arial", 14, "bold"),
            text_color="#1e3a5f"
        )
        selector_label.pack(side="left", padx=15, pady=10)
        
        self.month_var = ctk.StringVar(value="January")
        month_menu = ctk.CTkOptionMenu(
            selector_frame,
            values=["January", "February", "March", "April", "May", "June",
                    "July", "August", "September", "October", "November", "December"],
            variable=self.month_var,
            width=150,
            height=35,
            command=self.load_data
        )
        month_menu.pack(side="left", padx=10)
        
        self.year_var = ctk.StringVar(value=str(date.today().year))
        year_menu = ctk.CTkOptionMenu(
            selector_frame,
            values=[str(year) for year in range(2024, 2031)],
            variable=self.year_var,
            width=100,
            height=35,
            command=self.load_data
        )
        year_menu.pack(side="left", padx=10)
        
        # Refresh Button
        refresh_btn = ctk.CTkButton(
            selector_frame,
            text="Refresh Data",
            font=("Arial", 13, "bold"),
            fg_color="#3498db",
            hover_color="#2980b9",
            width=150,
            height=35,
            command=self.refresh_summary
        )
        refresh_btn.pack(side="right", padx=15)
        
        # ===== SUMMARY CARDS (5 Cards) =====
        summary_frame = ctk.CTkFrame(main_frame, fg_color="transparent")
        summary_frame.pack(fill="x", padx=20, pady=20)
        
        # Configure 5 columns
        for i in range(5):
            summary_frame.grid_columnconfigure(i, weight=1)
        
        # Card 1: Paid Amount
        paid_card = ctk.CTkFrame(summary_frame, fg_color="white", corner_radius=15)
        paid_card.grid(row=0, column=0, padx=5, pady=10, sticky="nsew")
        
        paid_label = ctk.CTkLabel(
            paid_card,
            text="Paid Amount",
            font=("Arial", 14),
            text_color="gray"
        )
        paid_label.pack(pady=(20, 5))
        
        self.paid_amount_label = ctk.CTkLabel(
            paid_card,
            text="Rs. 0",
            font=("Arial", 24, "bold"),
            text_color="#2ecc71"
        )
        self.paid_amount_label.pack(pady=(5, 20))
        
        # Card 2: Unpaid Amount
        unpaid_card = ctk.CTkFrame(summary_frame, fg_color="white", corner_radius=15)
        unpaid_card.grid(row=0, column=1, padx=5, pady=10, sticky="nsew")
        
        unpaid_label = ctk.CTkLabel(
            unpaid_card,
            text="Unpaid Amount",
            font=("Arial", 14),
            text_color="gray"
        )
        unpaid_label.pack(pady=(20, 5))
        
        self.unpaid_amount_label = ctk.CTkLabel(
            unpaid_card,
            text="Rs. 0",
            font=("Arial", 24, "bold"),
            text_color="#e74c3c"
        )
        self.unpaid_amount_label.pack(pady=(5, 20))
        
        # Card 3: Total Challans
        total_card = ctk.CTkFrame(summary_frame, fg_color="white", corner_radius=15)
        total_card.grid(row=0, column=2, padx=5, pady=10, sticky="nsew")
        
        total_label = ctk.CTkLabel(
            total_card,
            text="Total Challans",
            font=("Arial", 14),
            text_color="gray"
        )
        total_label.pack(pady=(20, 5))
        
        self.total_challans_label = ctk.CTkLabel(
            total_card,
            text="0",
            font=("Arial", 24, "bold"),
            text_color="#3498db"
        )
        self.total_challans_label.pack(pady=(5, 20))
        
        # Card 4: Total Amount of Challans
        total_amount_card = ctk.CTkFrame(summary_frame, fg_color="white", corner_radius=15)
        total_amount_card.grid(row=0, column=3, padx=5, pady=10, sticky="nsew")
        
        total_amount_label = ctk.CTkLabel(
            total_amount_card,
            text="Total Amount\nof Challans",
            font=("Arial", 14),
            text_color="gray"
        )
        total_amount_label.pack(pady=(20, 5))
        
        self.total_amount_label = ctk.CTkLabel(
            total_amount_card,
            text="Rs. 0",
            font=("Arial", 24, "bold"),
            text_color="#8e44ad"
        )
        self.total_amount_label.pack(pady=(5, 20))
        
        # Card 5: Total Arrears
        arrears_card = ctk.CTkFrame(summary_frame, fg_color="white", corner_radius=15)
        arrears_card.grid(row=0, column=4, padx=5, pady=10, sticky="nsew")
        
        arrears_label = ctk.CTkLabel(
            arrears_card,
            text="Total Arrears",
            font=("Arial", 14),
            text_color="gray"
        )
        arrears_label.pack(pady=(20, 5))
        
        self.total_arrears_label = ctk.CTkLabel(
            arrears_card,
            text="Rs. 0",
            font=("Arial", 24, "bold"),
            text_color="#f39c12"
        )
        self.total_arrears_label.pack(pady=(5, 20))
        
        # Table Section
        table_frame = ctk.CTkFrame(main_frame, fg_color="white", corner_radius=10)
        table_frame.pack(fill="both", expand=True, padx=20, pady=10)
        
        table_title = ctk.CTkLabel(
            table_frame,
            text="Payment Details",
            font=("Arial", 18, "bold"),
            text_color="#1e3a5f"
        )
        table_title.pack(pady=10)
        
        # Create custom table (SIMPLE, without canvas)
        self.create_table(table_frame)
        
        # Buttons
        button_frame = ctk.CTkFrame(main_frame, fg_color="transparent")
        button_frame.pack(fill="x", padx=20, pady=10)
        
        unpaid_btn = ctk.CTkButton(
            button_frame,
            text="Print Unpaid Students List (Excel)",
            font=("Arial", 14, "bold"),
            fg_color="#e74c3c",
            hover_color="#c0392b",
            width=300,
            height=45,
            command=self.print_unpaid
        )
        unpaid_btn.pack(side="left", padx=10)
        
        paid_btn = ctk.CTkButton(
            button_frame,
            text="Print Paid Student List (Excel)",
            font=("Arial", 14, "bold"),
            fg_color="#2ecc71",
            hover_color="#27ae60",
            width=300,
            height=45,
            command=self.print_paid
        )
        paid_btn.pack(side="left", padx=10)
        
        close_btn = ctk.CTkButton(
            main_frame,
            text="Close",
            font=("Arial", 13, "bold"),
            fg_color="#e74c3c",
            hover_color="#c0392b",
            width=100,
            height=40,
            command=self.destroy
        )
        close_btn.pack(pady=(0, 20))
    
    def create_table(self, parent):
        """Create a fixed-width table using Frames (SIMPLE, without canvas)"""
        
        # Define column widths (FIXED - no overlap possible)
        self.col_widths = [180, 160, 160, 90, 90, 90, 90, 90, 90, 120]  # Added Student Name
        self.headers = ["Bill ID", "Student Name", "Father Name", "Status", "Total Fees", "Paid", "Remaining", "Arrears", "Concession", "Payment Date"]
        
        # Table container
        table_container = ctk.CTkFrame(parent, fg_color="transparent")
        table_container.pack(fill="both", expand=True, padx=10, pady=10)
        
        # Create header row
        header_row = ctk.CTkFrame(table_container, fg_color="#1e3a5f", height=70)
        header_row.pack(fill="x", pady=(0, 2))
        header_row.pack_propagate(False)  # UI FIX: keep height consistent
        
        for i, (header_text, width) in enumerate(zip(self.headers, self.col_widths)):
            # Create a fixed-width frame for each header cell with padding
            cell_frame = ctk.CTkFrame(header_row, fg_color="#1e3a5f", width=width + 10, height=70)
            cell_frame.pack(side="left", padx=2, pady=0)
            cell_frame.pack_propagate(False)
            
            # Header label with wrapping
            header_label = ctk.CTkLabel(
                cell_frame,
                text=header_text,
                font=("Arial", 12, "bold"),
                text_color="white",
                anchor="center",
                justify="center",
                wraplength=width - 10
            )
            header_label.pack(fill="both", expand=True, padx=5, pady=10)
        header_row.pack_propagate(False)
        
        # Scrollable area for rows (Vertical only)
        self.scroll_frame = ctk.CTkScrollableFrame(
            table_container,
            fg_color="white",
            corner_radius=0
        )
        self.scroll_frame.pack(fill="both", expand=True, pady=(2, 0))
        
        # Store row references
        self.row_frames = []
    
    def load_data(self, *args):
        """Load payment data based on selected month and year"""
        
        # Clear existing rows
        for row_frame in self.row_frames:
            row_frame.destroy()
        self.row_frames = []
        
        # Get selected month
        month_name = self.month_var.get()
        year = int(self.year_var.get())
        
        # Get all challans
        challans = self.db.query(FeeChallan).all()
        
        total_paid = 0
        total_unpaid = 0
        total_challans = 0
        total_amount = 0
        total_arrears = 0
        total_concession = 0
        
        # Track seen bill IDs to avoid duplicates
        seen_bill_ids = set()
        
        for challan in challans:
            # Check if challan is for this month
            if not challan.challan_month or challan.challan_month.lower() != month_name.lower():
                continue
            
            # Skip if already seen (duplicate)
            if challan.bill_id in seen_bill_ids:
                continue
            seen_bill_ids.add(challan.bill_id)
            
            total_challans += 1
            total_amount += challan.amount_due
            total_arrears += challan.arrears
            total_concession += challan.fee_concession
            
            # Get student and guardian
            student = self.db.query(Student).filter(Student.id == challan.student_id).first()
            guardian = self.db.query(Guardian).filter(Guardian.family_id == challan.family_id).first()
            
            student_name = f"{student.first_name} {student.last_name}" if student else "N/A"
            father_name = guardian.guardian_name if guardian else "N/A"
            
            # Calculate amounts
            total_fee = challan.amount_due
            paid = challan.paid_amount
            remaining = challan.remaining_amount
            arrears = challan.arrears
            concession = challan.fee_concession
            payment_date = challan.payment_date
            
            # Add to totals
            total_paid += paid
            total_unpaid += remaining
            
            # Determine status
            if remaining <= 0:
                status = "Paid"
            else:
                status = "Unpaid"
            
            # Add to table
            self.add_row(challan.bill_id, student_name, father_name, status, total_fee, paid, remaining, arrears, concession, payment_date)
        
        # Update summary cards
        self.paid_amount_label.configure(text=f"Rs. {total_paid:.0f}")
        self.unpaid_amount_label.configure(text=f"Rs. {total_unpaid:.0f}")
        self.total_challans_label.configure(text=str(total_challans))
        self.total_amount_label.configure(text=f"Rs. {total_amount:.0f}")
        self.total_arrears_label.configure(text=f"Rs. {total_arrears:.0f}")
    
    def add_row(self, bill_id, student_name, father_name, status, total_fee, paid, remaining, arrears, concession, payment_date=None):
        """Add a row with fixed-width cells"""
        
        # Create row frame
        row_frame = ctk.CTkFrame(
            self.scroll_frame,
            fg_color="#f8f9fa" if len(self.row_frames) % 2 == 0 else "white",
            height=50
        )
        row_frame.pack(fill="x", pady=1)
        row_frame.pack_propagate(False)  # UI FIX: keep height consistent
        
        # Column 0: Bill ID
        cell_0 = ctk.CTkFrame(row_frame, fg_color="transparent", width=self.col_widths[0] + 10)
        cell_0.pack(side="left", padx=2)
        cell_0.pack_propagate(False)
        
        bill_label = ctk.CTkLabel(
            cell_0,
            text=bill_id,
            font=("Arial", 10),
            text_color="#2c3e50",
            anchor="w",
            wraplength=self.col_widths[0]
        )
        bill_label.pack(fill="both", expand=True, padx=5, pady=5)
        
        # Column 1: Student Name (NEW)
        cell_1 = ctk.CTkFrame(row_frame, fg_color="transparent", width=self.col_widths[1] + 10)
        cell_1.pack(side="left", padx=2)
        cell_1.pack_propagate(False)
        
        student_label = ctk.CTkLabel(
            cell_1,
            text=student_name,
            font=("Arial", 11),
            text_color="#2c3e50",
            anchor="w",
            wraplength=self.col_widths[1]
        )
        student_label.pack(fill="both", expand=True, padx=5, pady=5)
        
        # Column 2: Father Name
        cell_2 = ctk.CTkFrame(row_frame, fg_color="transparent", width=self.col_widths[2] + 10)
        cell_2.pack(side="left", padx=2)
        cell_2.pack_propagate(False)
        
        father_label = ctk.CTkLabel(
            cell_2,
            text=father_name,
            font=("Arial", 11),
            text_color="#2c3e50",
            anchor="w",
            wraplength=self.col_widths[2]
        )
        father_label.pack(fill="both", expand=True, padx=5, pady=5)
        
        # Column 3: Status
        cell_3 = ctk.CTkFrame(row_frame, fg_color="transparent", width=self.col_widths[3] + 10)
        cell_3.pack(side="left", padx=2)
        cell_3.pack_propagate(False)
        
        status_label = ctk.CTkLabel(
            cell_3,
            text=status,
            font=("Arial", 11, "bold"),
            text_color="#2ecc71" if status == "Paid" else "#e74c3c",
            anchor="center"
        )
        status_label.pack(fill="both", expand=True, padx=5, pady=5)
        
        # Column 4: Total Fees
        cell_4 = ctk.CTkFrame(row_frame, fg_color="transparent", width=self.col_widths[4] + 10)
        cell_4.pack(side="left", padx=2)
        cell_4.pack_propagate(False)
        
        total_label = ctk.CTkLabel(
            cell_4,
            text=f"Rs. {total_fee:.0f}",
            font=("Arial", 11),
            text_color="#2c3e50",
            anchor="center"
        )
        total_label.pack(fill="both", expand=True, padx=5, pady=5)
        
        # Column 5: Paid Amount
        cell_5 = ctk.CTkFrame(row_frame, fg_color="transparent", width=self.col_widths[5] + 10)
        cell_5.pack(side="left", padx=2)
        cell_5.pack_propagate(False)
        
        paid_label = ctk.CTkLabel(
            cell_5,
            text=f"Rs. {paid:.0f}",
            font=("Arial", 11),
            text_color="#2ecc71",
            anchor="center"
        )
        paid_label.pack(fill="both", expand=True, padx=5, pady=5)
        
        # Column 6: Remaining
        cell_6 = ctk.CTkFrame(row_frame, fg_color="transparent", width=self.col_widths[6] + 10)
        cell_6.pack(side="left", padx=2)
        cell_6.pack_propagate(False)
        
        remaining_label = ctk.CTkLabel(
            cell_6,
            text=f"Rs. {remaining:.0f}",
            font=("Arial", 11),
            text_color="#e74c3c",
            anchor="center"
        )
        remaining_label.pack(fill="both", expand=True, padx=5, pady=5)
        
        # Column 7: Arrears
        cell_7 = ctk.CTkFrame(row_frame, fg_color="transparent", width=self.col_widths[7] + 10)
        cell_7.pack(side="left", padx=2)
        cell_7.pack_propagate(False)
        
        arrears_label = ctk.CTkLabel(
            cell_7,
            text=f"Rs. {arrears:.0f}",
            font=("Arial", 11),
            text_color="#f39c12",
            anchor="center"
        )
        arrears_label.pack(fill="both", expand=True, padx=5, pady=5)
        
        # Column 8: Concession
        cell_8 = ctk.CTkFrame(row_frame, fg_color="transparent", width=self.col_widths[8] + 10)
        cell_8.pack(side="left", padx=2)
        cell_8.pack_propagate(False)
        
        concession_label = ctk.CTkLabel(
            cell_8,
            text=f"Rs. {concession:.0f}",
            font=("Arial", 11),
            text_color="#8e44ad",
            anchor="center"
        )
        concession_label.pack(fill="both", expand=True, padx=5, pady=5)
        
        # Column 9: Payment Date
        cell_9 = ctk.CTkFrame(row_frame, fg_color="transparent", width=self.col_widths[9] + 10)
        cell_9.pack(side="left", padx=2)
        cell_9.pack_propagate(False)
        
        # Format date if available
        if payment_date:
            date_text = payment_date.strftime("%Y-%m-%d")
        else:
            date_text = "-"
        
        payment_date_label = ctk.CTkLabel(
            cell_9,
            text=date_text,
            font=("Arial", 11),
            text_color="#3498db",
            anchor="center"
        )
        payment_date_label.pack(fill="both", expand=True, padx=5, pady=5)
        
        # Store row
        self.row_frames.append(row_frame)
    
    def print_unpaid(self):
        """Generate Excel file for unpaid students list"""
        try:
            month_name = self.month_var.get()
            year = int(self.year_var.get())
            
            challans = self.db.query(FeeChallan).filter(FeeChallan.remaining_amount > 0).all()
            
            if not challans:
                messagebox.showinfo("Info", "No unpaid challans found!")
                return
            
            wb = Workbook()
            ws = wb.active
            ws.title = f"Unpaid Students {month_name}"
            
            headers = ["Bill ID", "Student Name", "Father Name", "Class", "Month", "Total Fee", "Paid", "Remaining", "Arrears", "Concession", "Payment Date", "Status"]
            ws.append(headers)
            
            for cell in ws[1]:
                cell.font = cell.font.copy(bold=True)
                cell.fill = openpyxl.styles.PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
            
            for challan in challans:
                if challan.challan_month.lower() != month_name.lower():
                    continue
                
                student = self.db.query(Student).filter(Student.id == challan.student_id).first()
                guardian = self.db.query(Guardian).filter(Guardian.family_id == challan.family_id).first()
                
                payment_date_text = challan.payment_date.strftime("%Y-%m-%d") if challan.payment_date else "N/A"
                
                ws.append([
                    challan.bill_id,
                    f"{student.first_name} {student.last_name}" if student else "N/A",
                    guardian.guardian_name if guardian else "N/A",
                    student.class_grade if student else "N/A",
                    challan.challan_month,
                    challan.amount_due,
                    challan.paid_amount,
                    challan.remaining_amount,
                    challan.arrears,
                    challan.fee_concession,
                    payment_date_text,
                    "Unpaid"
                ])
            
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 30)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            filename = f"Unpaid_Students_{month_name}_{year}.xlsx"
            filepath = os.path.join(self.reports_dir, filename)
            wb.save(filepath)
            
            messagebox.showinfo("Success", f"Unpaid Students List saved to: {filepath}")
            
        except Exception as e:
            messagebox.showerror("Error", f"Failed to print list: {str(e)}")
    
    def print_paid(self):
        """Generate Excel file for paid students list"""
        try:
            month_name = self.month_var.get()
            year = int(self.year_var.get())
            
            challans = self.db.query(FeeChallan).filter(FeeChallan.remaining_amount <= 0).all()
            
            if not challans:
                messagebox.showinfo("Info", "No paid challans found!")
                return
            
            wb = Workbook()
            ws = wb.active
            ws.title = f"Paid Students {month_name}"
            
            headers = ["Bill ID", "Student Name", "Father Name", "Class", "Month", "Total Fee", "Paid", "Remaining", "Arrears", "Concession", "Payment Date", "Status", "Receipt No"]
            ws.append(headers)
            
            for cell in ws[1]:
                cell.font = cell.font.copy(bold=True)
                cell.fill = openpyxl.styles.PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
            
            for challan in challans:
                if challan.challan_month.lower() != month_name.lower():
                    continue
                
                student = self.db.query(Student).filter(Student.id == challan.student_id).first()
                guardian = self.db.query(Guardian).filter(Guardian.family_id == challan.family_id).first()
                
                payment_date_text = challan.payment_date.strftime("%Y-%m-%d") if challan.payment_date else "N/A"
                
                ws.append([
                    challan.bill_id,
                    f"{student.first_name} {student.last_name}" if student else "N/A",
                    guardian.guardian_name if guardian else "N/A",
                    student.class_grade if student else "N/A",
                    challan.challan_month,
                    challan.amount_due,
                    challan.paid_amount,
                    challan.remaining_amount,
                    challan.arrears,
                    challan.fee_concession,
                    payment_date_text,
                    "Paid",
                    challan.receipt_number if challan.receipt_number else "N/A"
                ])
            
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 30)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            filename = f"Paid_Students_{month_name}_{year}.xlsx"
            filepath = os.path.join(self.reports_dir, filename)
            wb.save(filepath)
            
            messagebox.showinfo("Success", f"Paid Students List saved to: {filepath}")
            
        except Exception as e:
            messagebox.showerror("Error", f"Failed to print list: {str(e)}")